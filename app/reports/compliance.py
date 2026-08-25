from __future__ import annotations

from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..repositories.compliance import ComplianceDataset


def _safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _write_sheet(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, len(headers) + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(map(len, values)) + 2, 48)


def render_compliance_xlsx(dataset: ComplianceDataset) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_rows = [
        ["Контур", "Внешний" if dataset.scope == "internet" else "Организация"],
        ["Дата оценки", dataset.assessment_date.isoformat()],
        ["Допустимый возраст сканирования, дней", dataset.summary["freshness_days"]],
        ["Активов всего", dataset.summary["assets_total"]],
        ["Со свежим сканированием", dataset.summary["fresh_assets"]],
        ["Не соответствует свежести", dataset.summary["stale_assets"]],
        ["Активов с критическими уязвимостями", dataset.summary["affected_assets"]],
        ["Критических находок", dataset.summary["critical_findings"]],
    ]
    _write_sheet(summary_sheet, ["Показатель", "Значение"], summary_rows)

    findings = workbook.create_sheet("Критические")
    _write_sheet(findings, ["Актив", "Имя", "IP", "Тип актива", "Категория", "Дата сканирования", "CVE/ID", "Уязвимость", "CVSS", "Источник"], [
        [row.get("asset_id"), row.get("display_name"), row.get("ip_address"), row.get("asset_type"), row.get("asset_category"), row.get("scan_at"), row.get("cve") or row.get("vulnerability_id"), row.get("vulnerability_name"), row.get("cvss_score"), row.get("source_type")]
        for row in dataset.findings
    ])

    stale = workbook.create_sheet("Не соответствует свежести")
    _write_sheet(stale, ["Актив", "Имя", "IP", "Тип актива", "Категория", "Дата сканирования", "Возраст, дней", "Причина"], [
        [row.get("asset_id"), row.get("display_name"), row.get("ip_address"), row.get("asset_type"), row.get("asset_category"), row.get("scan_at"), row.get("age_days"), str(row.get("freshness_reason") or "")]
        for row in dataset.stale_assets
    ])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_font() -> str:
    candidates = (
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    )
    for path in candidates:
        if path.exists():
            pdfmetrics.registerFont(TTFont("ComplianceFont", str(path)))
            return "ComplianceFont"
    return "Helvetica"


def _paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape("" if value is None else str(value)), style)


def render_compliance_pdf(dataset: ComplianceDataset) -> bytes:
    output = BytesIO()
    font = _pdf_font()
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName=font, fontSize=16, leading=20, alignment=TA_CENTER, spaceAfter=8 * mm)
    body = ParagraphStyle("ReportBody", parent=styles["BodyText"], fontName=font, fontSize=7.5, leading=9)
    heading = ParagraphStyle("ReportHeading", parent=styles["Heading2"], fontName=font, fontSize=12, leading=15, spaceBefore=4 * mm, spaceAfter=3 * mm)
    document = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    story = [
        Paragraph("Отчет по критическим уязвимостям", title),
        Paragraph(f"Контур: {'внешний (10.255.0.0/16)' if dataset.scope == 'internet' else 'пользовательские устройства и серверы организации'}", body),
        Paragraph(f"Дата оценки: {dataset.assessment_date.isoformat()}. Допустимый возраст результатов: 30 календарных дней.", body),
        Spacer(1, 4 * mm),
    ]
    summary_data = [[_paragraph("Показатель", body), _paragraph("Значение", body)]] + [
        [_paragraph(label, body), _paragraph(dataset.summary[key], body)]
        for label, key in (
            ("Активов всего", "assets_total"), ("Со свежим сканированием", "fresh_assets"),
            ("Не соответствует свежести", "stale_assets"), ("Активов с критическими уязвимостями", "affected_assets"),
            ("Критических находок", "critical_findings"),
        )
    ]
    summary = Table(summary_data, colWidths=[95 * mm, 35 * mm], repeatRows=1)
    common = [("FONTNAME", (0, 0), (-1, -1), font), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#AAB7C4")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F8")])]
    summary.setStyle(TableStyle(common))
    story.extend([summary, Paragraph("Критические уязвимости", heading)])
    finding_headers = ["Актив", "IP", "Тип", "Дата сканирования", "CVE/ID", "Уязвимость", "CVSS"]
    finding_rows = [[row.get("asset_id"), row.get("ip_address"), row.get("asset_type"), row.get("scan_at"), row.get("cve") or row.get("vulnerability_id"), row.get("vulnerability_name"), row.get("cvss_score")] for row in dataset.findings]
    table = Table([[ _paragraph(value, body) for value in finding_headers]] + [[_paragraph(value, body) for value in row] for row in finding_rows], colWidths=[34*mm, 26*mm, 28*mm, 42*mm, 32*mm, 80*mm, 15*mm], repeatRows=1)
    table.setStyle(TableStyle(common))
    story.append(table)
    story.extend([PageBreak(), Paragraph("Не соответствует требованиям по свежести", heading)])
    stale_headers = ["Актив", "IP", "Тип", "Категория", "Дата сканирования", "Возраст, дней", "Причина"]
    stale_rows = [[row.get("asset_id"), row.get("ip_address"), row.get("asset_type"), row.get("asset_category"), row.get("scan_at"), row.get("age_days"), row.get("freshness_reason")] for row in dataset.stale_assets]
    stale_table = Table([[_paragraph(value, body) for value in stale_headers]] + [[_paragraph(value, body) for value in row] for row in stale_rows], colWidths=[38*mm, 28*mm, 35*mm, 34*mm, 45*mm, 25*mm, 50*mm], repeatRows=1)
    stale_table.setStyle(TableStyle(common))
    story.append(stale_table)
    document.build(story)
    return output.getvalue()
