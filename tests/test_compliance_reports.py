from datetime import UTC, date, datetime
from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from app.reports.compliance import render_compliance_pdf, render_compliance_xlsx
from app.repositories.compliance import ComplianceDataset


def _dataset():
    asset = {
        "asset_id": "asset-1",
        "display_name": "Рабочая станция",
        "ip_address": "10.255.1.2",
        "asset_type": "workstation",
        "asset_category": "user_device",
        "scan_at": "2026-08-20T10:00:00+00:00",
        "age_days": 5,
        "freshness_reason": "fresh",
    }
    finding = {
        **asset,
        "cve": "=HYPERLINK(\"bad\")",
        "vulnerability_id": "v-1",
        "vulnerability_name": "Критическая уязвимость",
        "severity": "critical",
        "cvss_score": 9.8,
        "source_type": "scanner",
    }
    stale = {**asset, "asset_id": "asset-2", "is_fresh": False, "age_days": 31, "freshness_reason": "too_old"}
    return ComplianceDataset(
        scope="internet",
        assessment_date=date(2026, 8, 25),
        generated_at=datetime(2026, 8, 25, tzinfo=UTC),
        summary={
            "scope": "internet", "assessment_date": "2026-08-25", "freshness_days": 30,
            "assets_total": 2, "fresh_assets": 1, "stale_assets": 1,
            "affected_assets": 1, "critical_findings": 1, "unique_vulnerabilities": 1,
            "oldest_fresh_scan_at": asset["scan_at"], "freshest_scan_at": asset["scan_at"],
            "by_asset_category": {"user_device": 1, "server": 0, "unclassified": 1},
        },
        findings=(finding,), assets=(asset, stale), stale_assets=(stale,), diagnostics={},
    )


def test_xlsx_report_has_required_sheets_and_escapes_formulas():
    workbook = load_workbook(BytesIO(render_compliance_xlsx(_dataset())))
    assert workbook.sheetnames == ["Сводка", "Критические", "Не соответствует свежести"]
    assert workbook["Критические"]["G2"].value.startswith("'")


def test_pdf_report_is_readable_and_contains_sections():
    data = render_compliance_pdf(_dataset())
    assert data.startswith(b"%PDF")
    reader = PdfReader(BytesIO(data))
    assert len(reader.pages) >= 1
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Отчет по критическим уязвимостям" in text
    assert "Не соответствует требованиям по свежести" in text
    assert "Критическая уязвимость" in text
